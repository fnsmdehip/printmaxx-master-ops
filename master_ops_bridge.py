#!/usr/bin/env python3
"""
MASTER OPS BRIDGE — Shared module for xlsx intelligence
========================================================
Parses PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx into cached JSON.
Any automation script imports this for ops data.

Usage as module:
    from master_ops_bridge import MasterOpsBridge
    bridge = MasterOpsBridge()
    ops = bridge.get_all_ops()
    priority = bridge.get_priority_launch()
    synergy = bridge.get_synergy_stacks()

Usage as CLI:
    python3 AUTOMATIONS/master_ops_bridge.py --rebuild    # Force rebuild cache
    python3 AUTOMATIONS/master_ops_bridge.py --stats      # Print cache stats
    python3 AUTOMATIONS/master_ops_bridge.py --query CONTENT  # Ops by category
    python3 AUTOMATIONS/master_ops_bridge.py --venture C01    # Venture details
    python3 AUTOMATIONS/master_ops_bridge.py --synergy        # Synergy stacks
    python3 AUTOMATIONS/master_ops_bridge.py --blockers       # Blocker summary
    python3 AUTOMATIONS/master_ops_bridge.py --json           # Full cache as JSON (for piping)
"""

from __future__ import annotations

import argparse
import json
import sys
import threading
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent

XLSX_PATTERNS = [
    "PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx",
    "PRINTMAXX_MASTER_OPS*.xlsx",
]

# Sheets we parse (structured data with real headers).
# Skipped: DEEP PLAYBOOK (unstructured 2999 rows), LLM ALPHA THESIS (unstructured 79 rows),
#           NSFW COMPLIANCE, BROWSER & PROXY STACK, SYSTEM_EVIDENCE.
PARSED_SHEETS = [
    "ALL OPS MASTER",
    "PRIORITY LAUNCH",
    "SYNERGY STACKS",
    "AUTO_STATUS_LIVE",
    "VENTURE_AUTOMATION_MAP",
    "ALPHA_THESIS_INDEX",
    "VIDEO & MEDIA STACK",
    "HOSTING & DEPLOY",
    "LEAD GEN STACK",
    "EXISTING INFRA",
    "DEEP_PLAYBOOK_INDEX",
    "RBI SYSTEM",
    "ETC_EXPANSION_QUEUE",
    "PRIORITY_AUTOMATION_EXEC",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_str(val: Any) -> str:
    """Convert a cell value to a stripped string, handling None."""
    if val is None:
        return ""
    return str(val).strip()


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Best-effort float conversion."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    """Best-effort int conversion."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


# ---------------------------------------------------------------------------
# MasterOpsBridge
# ---------------------------------------------------------------------------

class MasterOpsBridge:
    """Shared interface to Master Ops xlsx data.

    On first instantiation (or when the cache is stale / force_rebuild=True),
    the xlsx is parsed via openpyxl and written to a JSON cache file.
    Subsequent instantiations read from the cache, which is thread-safe.
    """

    CACHE_PATH = PROJECT_ROOT / "AUTOMATIONS" / "master_ops_cache.json"
    CACHE_MAX_AGE_HOURS = 12  # Rebuild if cache older than this

    _lock = threading.Lock()

    def __init__(self, force_rebuild: bool = False) -> None:
        self._data: Dict[str, Any] = {}
        if force_rebuild or self._cache_stale():
            self._rebuild_cache()
        self._load_cache()

    # ------------------------------------------------------------------
    # Cache management
    # ------------------------------------------------------------------

    def _find_latest_xlsx(self) -> Optional[Path]:
        """Find latest PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx in project root."""
        for pattern in XLSX_PATTERNS:
            matches = sorted(PROJECT_ROOT.glob(pattern), key=lambda p: p.name)
            if matches:
                return matches[-1]
        return None

    def _cache_stale(self) -> bool:
        """Check if cache needs rebuild."""
        if not self.CACHE_PATH.exists():
            return True
        try:
            age_hours = (time.time() - self.CACHE_PATH.stat().st_mtime) / 3600
            return age_hours > self.CACHE_MAX_AGE_HOURS
        except OSError:
            return True

    def _rebuild_cache(self) -> None:
        """Parse xlsx and write JSON cache atomically."""
        xlsx_path = self._find_latest_xlsx()
        if not xlsx_path:
            print("[master_ops_bridge] No master xlsx found in project root", file=sys.stderr)
            return

        try:
            import openpyxl  # noqa: delayed import — may not be installed
        except ImportError:
            print("[master_ops_bridge] openpyxl not installed — run: pip3 install openpyxl", file=sys.stderr)
            return

        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        except Exception as exc:
            print("[master_ops_bridge] Failed to open xlsx: {}".format(exc), file=sys.stderr)
            return

        cache: Dict[str, Any] = {
            "source": xlsx_path.name,
            "generated_at": datetime.now().isoformat(),
            "sheets": {},
        }

        for sheet_name in PARSED_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                cache["sheets"][sheet_name] = []
                continue

            # Build header list — handle None headers gracefully
            headers: List[str] = []
            for i, h in enumerate(rows[0]):
                headers.append(str(h).strip() if h else "col_{}".format(i))

            sheet_data: List[Dict[str, Any]] = []
            for row in rows[1:]:
                entry: Dict[str, Any] = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        # Convert datetimes to ISO strings for JSON serialization
                        if isinstance(val, datetime):
                            entry[headers[i]] = val.isoformat()
                        else:
                            entry[headers[i]] = val
                # Skip entirely empty rows
                if any(v is not None and v != "" for v in entry.values()):
                    sheet_data.append(entry)

            cache["sheets"][sheet_name] = sheet_data

        wb.close()

        # Atomic write: write to .tmp then rename
        with self._lock:
            tmp = self.CACHE_PATH.with_suffix(".tmp")
            try:
                with open(tmp, "w") as f:
                    json.dump(cache, f, indent=2, default=str)
                tmp.rename(self.CACHE_PATH)
            except Exception as exc:
                print("[master_ops_bridge] Cache write failed: {}".format(exc), file=sys.stderr)
                if tmp.exists():
                    tmp.unlink()

        rows_total = sum(len(v) for v in cache["sheets"].values())
        print("[master_ops_bridge] Cache rebuilt from {}: {} sheets, {} rows".format(
            xlsx_path.name, len(cache["sheets"]), rows_total))

    def _load_cache(self) -> None:
        """Load JSON cache into memory (thread-safe read)."""
        with self._lock:
            if self.CACHE_PATH.exists():
                try:
                    with open(self.CACHE_PATH) as f:
                        self._data = json.load(f)
                except (json.JSONDecodeError, OSError) as exc:
                    print("[master_ops_bridge] Cache read failed: {}".format(exc), file=sys.stderr)
                    self._data = {}
            else:
                self._data = {}

    def _sheet(self, name: str) -> List[Dict[str, Any]]:
        """Get a parsed sheet by name from cached data."""
        sheets = self._data.get("sheets", {})
        return sheets.get(name, [])

    # ------------------------------------------------------------------
    # Core getters
    # ------------------------------------------------------------------

    def get_all_ops(self) -> List[Dict[str, Any]]:
        """All 182 ops from ALL OPS MASTER sheet."""
        return self._sheet("ALL OPS MASTER")

    def get_ops_by_category(self, category: str) -> List[Dict[str, Any]]:
        """Ops filtered by category (CONTENT, SERVICE, APP, etc.)."""
        cat_upper = category.upper()
        return [op for op in self.get_all_ops()
                if _safe_str(op.get("CATEGORY")).upper() == cat_upper]

    def get_op(self, op_id: str) -> Optional[Dict[str, Any]]:
        """Single op by ID (C01, S01, N68, etc.)."""
        op_upper = op_id.upper()
        for op in self.get_all_ops():
            if _safe_str(op.get("OP_ID")).upper() == op_upper:
                return op
        return None

    # ------------------------------------------------------------------
    # Priority & Readiness
    # ------------------------------------------------------------------

    def get_priority_launch(self) -> List[Dict[str, Any]]:
        """From PRIORITY LAUNCH sheet — ranked ops with time-to-first-$."""
        return self._sheet("PRIORITY LAUNCH")

    def get_readiness_summary(self) -> Dict[str, int]:
        """From AUTO_STATUS_LIVE — counts of READY/BUILD/BLOCKED."""
        counts: Dict[str, int] = defaultdict(int)
        for row in self._sheet("AUTO_STATUS_LIVE"):
            readiness = _safe_str(row.get("READINESS")).upper()
            if readiness:
                counts[readiness] += 1
        return dict(counts)

    def get_ready_ops(self) -> List[Dict[str, Any]]:
        """Only ops with readiness=READY from AUTO_STATUS_LIVE."""
        return [row for row in self._sheet("AUTO_STATUS_LIVE")
                if _safe_str(row.get("READINESS")).upper() == "READY"]

    def get_blocked_ops(self) -> List[Dict[str, Any]]:
        """Only ops with blockers from AUTO_STATUS_LIVE."""
        return [row for row in self._sheet("AUTO_STATUS_LIVE")
                if _safe_str(row.get("BLOCKER_KEY"))]

    # ------------------------------------------------------------------
    # Synergy & Compounding
    # ------------------------------------------------------------------

    def get_synergy_stacks(self) -> List[Dict[str, Any]]:
        """From SYNERGY STACKS — all 26 synergy combos with multipliers."""
        return self._sheet("SYNERGY STACKS")

    def get_synergy_for_op(self, op_id: str) -> List[Dict[str, Any]]:
        """Find synergy stacks that include a specific op."""
        op_upper = op_id.upper()
        results = []
        for stack in self.get_synergy_stacks():
            methods_combined = _safe_str(stack.get("METHODS_COMBINED")).upper()
            if op_upper in methods_combined:
                results.append(stack)
        return results

    def get_top_synergies(self, n: int = 5) -> List[Dict[str, Any]]:
        """Top N synergy stacks by score."""
        stacks = self.get_synergy_stacks()
        return sorted(stacks, key=lambda s: _safe_float(s.get("SYNERGY_SCORE")), reverse=True)[:n]

    # ------------------------------------------------------------------
    # Venture Mapping
    # ------------------------------------------------------------------

    def get_venture_map(self) -> List[Dict[str, Any]]:
        """From VENTURE_AUTOMATION_MAP — all venture automation mappings."""
        return self._sheet("VENTURE_AUTOMATION_MAP")

    def get_venture(self, venture_id: str) -> Optional[Dict[str, Any]]:
        """Single venture by ID."""
        vid_upper = venture_id.upper()
        for v in self.get_venture_map():
            if _safe_str(v.get("VENTURE_ID")).upper() == vid_upper:
                return v
        return None

    def get_ventures_by_lane(self, lane: str) -> List[Dict[str, Any]]:
        """Ventures filtered by lane (app_factory, freelance_arbitrage, etc.)."""
        lane_upper = lane.upper()
        return [v for v in self.get_venture_map()
                if _safe_str(v.get("LANE")).upper() == lane_upper]

    def get_ventures_by_blocker(self, blocker_key: str) -> List[Dict[str, Any]]:
        """Ventures blocked by a specific blocker key."""
        bk_upper = blocker_key.upper()
        return [v for v in self.get_venture_map()
                if _safe_str(v.get("BLOCKER_KEY")).upper() == bk_upper]

    # ------------------------------------------------------------------
    # Alpha Thesis
    # ------------------------------------------------------------------

    def get_alpha_theses(self) -> List[Dict[str, Any]]:
        """From ALPHA_THESIS_INDEX — all alpha opportunities with edge durations."""
        return self._sheet("ALPHA_THESIS_INDEX")

    def get_alpha_by_lane(self, lane: str) -> List[Dict[str, Any]]:
        """Alpha theses filtered by lane."""
        lane_upper = lane.upper()
        return [a for a in self.get_alpha_theses()
                if _safe_str(a.get("LANE")).upper() == lane_upper]

    # ------------------------------------------------------------------
    # Tool Stacks
    # ------------------------------------------------------------------

    _TOOL_SHEET_MAP = {
        "video": "VIDEO & MEDIA STACK",
        "hosting": "HOSTING & DEPLOY",
        "lead_gen": "LEAD GEN STACK",
    }

    def get_tool_stack(self, stack_type: str) -> List[Dict[str, Any]]:
        """Get tools by type: 'video', 'hosting', 'lead_gen'."""
        sheet_name = self._TOOL_SHEET_MAP.get(stack_type.lower(), "")
        if not sheet_name:
            return []
        return self._sheet(sheet_name)

    def get_all_tool_stacks(self) -> Dict[str, List[Dict[str, Any]]]:
        """All tool stacks grouped by type."""
        return {key: self._sheet(sheet) for key, sheet in self._TOOL_SHEET_MAP.items()}

    # ------------------------------------------------------------------
    # Infrastructure
    # ------------------------------------------------------------------

    def get_existing_infra(self) -> List[Dict[str, Any]]:
        """From EXISTING INFRA — what's already built."""
        return self._sheet("EXISTING INFRA")

    def get_infra_by_category(self, category: str) -> List[Dict[str, Any]]:
        """Infrastructure filtered by category (SCRAPER, PIPELINE, etc.)."""
        cat_upper = category.upper()
        return [item for item in self.get_existing_infra()
                if _safe_str(item.get("CATEGORY")).upper() == cat_upper]

    # ------------------------------------------------------------------
    # Playbooks
    # ------------------------------------------------------------------

    def get_playbook_index(self) -> List[Dict[str, Any]]:
        """From DEEP_PLAYBOOK_INDEX — indexed playbook steps for all 37 ops."""
        return self._sheet("DEEP_PLAYBOOK_INDEX")

    def get_playbook_for_op(self, op_id: str) -> List[Dict[str, Any]]:
        """Get playbook steps for a specific op."""
        op_upper = op_id.upper()
        return [step for step in self.get_playbook_index()
                if _safe_str(step.get("PLAYBOOK_OP")).upper() == op_upper]

    # ------------------------------------------------------------------
    # RBI System
    # ------------------------------------------------------------------

    def get_rbi_system(self) -> List[Dict[str, Any]]:
        """From RBI SYSTEM — recurring build inspection audits."""
        return self._sheet("RBI SYSTEM")

    # ------------------------------------------------------------------
    # Expansion Queue
    # ------------------------------------------------------------------

    def get_expansion_queue(self) -> List[Dict[str, Any]]:
        """From ETC_EXPANSION_QUEUE — ops queued for scaling."""
        return self._sheet("ETC_EXPANSION_QUEUE")

    # ------------------------------------------------------------------
    # Priority Automation Exec
    # ------------------------------------------------------------------

    def get_priority_automation_exec(self) -> List[Dict[str, Any]]:
        """From PRIORITY_AUTOMATION_EXEC — top ops for automated execution."""
        return self._sheet("PRIORITY_AUTOMATION_EXEC")

    # ------------------------------------------------------------------
    # Blockers
    # ------------------------------------------------------------------

    def get_blocker_summary(self) -> List[Dict[str, Any]]:
        """Aggregate blockers: which blocker keys block which ventures/ops, counts."""
        blocker_map: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            "blocker_key": "",
            "blocked_ops": [],
            "blocked_ventures": [],
            "op_count": 0,
            "venture_count": 0,
        })

        # Blocked ops from AUTO_STATUS_LIVE
        for row in self._sheet("AUTO_STATUS_LIVE"):
            bk = _safe_str(row.get("BLOCKER_KEY"))
            if bk:
                entry = blocker_map[bk]
                entry["blocker_key"] = bk
                op_id = _safe_str(row.get("OP_ID"))
                if op_id and op_id not in entry["blocked_ops"]:
                    entry["blocked_ops"].append(op_id)
                    entry["op_count"] = len(entry["blocked_ops"])

        # Blocked ventures from VENTURE_AUTOMATION_MAP
        for row in self._sheet("VENTURE_AUTOMATION_MAP"):
            bk = _safe_str(row.get("BLOCKER_KEY"))
            if bk:
                entry = blocker_map[bk]
                entry["blocker_key"] = bk
                vid = _safe_str(row.get("VENTURE_ID"))
                if vid and vid not in entry["blocked_ventures"]:
                    entry["blocked_ventures"].append(vid)
                    entry["venture_count"] = len(entry["blocked_ventures"])

        # Sort by total impact descending
        results = sorted(blocker_map.values(),
                         key=lambda b: b["op_count"] + b["venture_count"],
                         reverse=True)
        return results

    def get_blocker_impact(self, blocker_key: str) -> Dict[str, Any]:
        """How many ops/ventures a specific blocker affects."""
        bk_upper = blocker_key.upper()
        for b in self.get_blocker_summary():
            if _safe_str(b.get("blocker_key")).upper() == bk_upper:
                return b
        return {"blocker_key": blocker_key, "blocked_ops": [], "blocked_ventures": [],
                "op_count": 0, "venture_count": 0}

    # ------------------------------------------------------------------
    # Composite Intelligence
    # ------------------------------------------------------------------

    def get_ops_intelligence_brief(self, venture_type: str) -> Dict[str, Any]:
        """Full intelligence brief for a venture type — combines ops, synergy, alpha, tools, blockers.

        This is the main method that intelligence_router.py and agent_swarm.py
        should call to get a comprehensive picture of a venture's state.
        """
        vt_upper = venture_type.upper()

        # Ops in this category
        category_ops = self.get_ops_by_category(vt_upper)
        op_ids = [_safe_str(op.get("OP_ID")) for op in category_ops]

        # Ready ops for this category
        ready = [op for op in self.get_ready_ops()
                 if _safe_str(op.get("CATEGORY")).upper() == vt_upper]

        # Synergies involving any op in this category
        synergies = []
        seen_packages: set = set()
        for oid in op_ids:
            for s in self.get_synergy_for_op(oid):
                pid = _safe_str(s.get("PACKAGE_ID"))
                if pid not in seen_packages:
                    seen_packages.add(pid)
                    synergies.append(s)

        # Alpha theses with matching lane
        alpha_theses = self.get_alpha_by_lane(vt_upper)

        # Tool stacks — include all (agents choose what they need)
        tools = self.get_all_tool_stacks()

        # Blockers affecting ops in this category
        blocked = [op for op in self.get_blocked_ops()
                   if _safe_str(op.get("CATEGORY")).upper() == vt_upper]

        # Playbook availability
        playbook_ops = []
        for oid in op_ids:
            steps = self.get_playbook_for_op(oid)
            if steps:
                playbook_ops.append({"op_id": oid, "step_count": len(steps)})

        # Priority launches for this category
        priority = [p for p in self.get_priority_launch()
                    if _safe_str(p.get("OP_ID")) in op_ids]

        return {
            "venture_type": vt_upper,
            "ops": category_ops,
            "ops_count": len(category_ops),
            "ready_ops": ready,
            "ready_count": len(ready),
            "synergies": synergies,
            "alpha_theses": alpha_theses,
            "tools": tools,
            "blockers": blocked,
            "blocker_count": len(blocked),
            "playbook_available": playbook_ops,
            "priority_launches": priority,
        }

    # ------------------------------------------------------------------
    # Decision engine support
    # ------------------------------------------------------------------

    def get_decision_weights(self, op_id: str) -> Dict[str, Any]:
        """Get scoring weights for an op — automation_score, signal_count, readiness, synergy potential."""
        op_upper = op_id.upper()

        # Base info from AUTO_STATUS_LIVE
        status_row: Optional[Dict[str, Any]] = None
        for row in self._sheet("AUTO_STATUS_LIVE"):
            if _safe_str(row.get("OP_ID")).upper() == op_upper:
                status_row = row
                break

        if not status_row:
            return {
                "op_id": op_id,
                "found": False,
                "automation_score": 0,
                "signal_count": 0,
                "readiness": "UNKNOWN",
                "synergy_potential": 0,
            }

        synergy_stacks = self.get_synergy_for_op(op_upper)
        max_synergy_score = max(
            (_safe_float(s.get("SYNERGY_SCORE")) for s in synergy_stacks),
            default=0.0,
        )

        return {
            "op_id": op_upper,
            "found": True,
            "automation_score": _safe_float(status_row.get("AUTOMATION_SCORE_100")),
            "signal_count": _safe_int(status_row.get("SIGNAL_COUNT")),
            "readiness": _safe_str(status_row.get("READINESS")),
            "blocker_key": _safe_str(status_row.get("BLOCKER_KEY")),
            "approval_ok": _safe_str(status_row.get("APPROVAL_OK")),
            "expected_profit_hint": _safe_str(status_row.get("EXPECTED_PROFIT_HINT_USD")),
            "synergy_potential": max_synergy_score,
            "synergy_stacks_count": len(synergy_stacks),
        }

    # ------------------------------------------------------------------
    # Daily tactical engine support
    # ------------------------------------------------------------------

    def get_daily_actionable(self) -> List[Dict[str, Any]]:
        """Ops that are READY + have commands + not blocked. Sorted by priority rank.

        Merges data from AUTO_STATUS_LIVE, VENTURE_AUTOMATION_MAP (for commands),
        and PRIORITY_AUTOMATION_EXEC (for rank).
        """
        # Build command lookup from VENTURE_AUTOMATION_MAP
        cmd_lookup: Dict[str, str] = {}
        for v in self.get_venture_map():
            vid = _safe_str(v.get("VENTURE_ID")).upper()
            cmd = _safe_str(v.get("COMMAND_TEMPLATE"))
            if vid and cmd:
                cmd_lookup[vid] = cmd

        # Build rank lookup from PRIORITY_AUTOMATION_EXEC
        rank_lookup: Dict[str, int] = {}
        for p in self.get_priority_automation_exec():
            oid = _safe_str(p.get("OP_ID")).upper()
            rank = _safe_int(p.get("RANK"), 9999)
            if oid:
                rank_lookup[oid] = rank

        # Human-only blockers — ops with these can still run automation parts
        HUMAN_BLOCKERS = {
            "GUMROAD_ACCOUNT", "STORE_ACCOUNT_AND_PAYMENT",
            "X_MULTI_ACCOUNT_STACK", "EMAIL_INFRA",
            "FIVERR_UPWORK_ACCOUNT", "ACCOUNT_EBAY",
            "ACCOUNT_ETSY", "ACCOUNT_AMAZON",
        }

        actionable: List[Dict[str, Any]] = []
        for row in self.get_ready_ops():
            op_id = _safe_str(row.get("OP_ID")).upper()
            blocker = _safe_str(row.get("BLOCKER_KEY")).upper()
            next_action = _safe_str(row.get("NEXT_AUTOMATION_ACTION"))
            cmd = cmd_lookup.get(op_id, "")
            rank = rank_lookup.get(op_id, 9999)

            # Include ops even with human-only blockers (mark them)
            is_blocked = bool(blocker) and blocker not in HUMAN_BLOCKERS

            actionable.append({
                "op_id": op_id,
                "op_name": _safe_str(row.get("OP_NAME")),
                "category": _safe_str(row.get("CATEGORY")),
                "lane": _safe_str(row.get("LANE")),
                "automation_score": _safe_float(row.get("AUTOMATION_SCORE_100")),
                "signal_count": _safe_int(row.get("SIGNAL_COUNT")),
                "next_action": next_action,
                "command_template": cmd,
                "priority_rank": rank,
                "blocker": blocker if blocker else "",
                "human_blocked": blocker in HUMAN_BLOCKERS,
                "fully_blocked": is_blocked,
            })

        actionable.sort(key=lambda x: x["priority_rank"])
        return actionable

    # ------------------------------------------------------------------
    # Loop closer support
    # ------------------------------------------------------------------

    def get_unblocked_since(self, hours: int = 24) -> List[Dict[str, Any]]:
        """Ops that changed from BLOCKED to READY recently.

        Compares current cache with a previous snapshot if available.
        Falls back to empty list if no prior snapshot exists.
        """
        previous_path = self.CACHE_PATH.with_suffix(".prev.json")
        current_ready_ids = {
            _safe_str(r.get("OP_ID")).upper()
            for r in self.get_ready_ops()
            if not _safe_str(r.get("BLOCKER_KEY"))
        }

        if not previous_path.exists():
            # No prior snapshot — return empty (first run)
            return []

        try:
            with open(previous_path) as f:
                prev = json.load(f)
        except (json.JSONDecodeError, OSError):
            return []

        prev_blocked_ids: set = set()
        for row in prev.get("sheets", {}).get("AUTO_STATUS_LIVE", []):
            if _safe_str(row.get("BLOCKER_KEY")):
                prev_blocked_ids.add(_safe_str(row.get("OP_ID")).upper())

        # Ops that WERE blocked but are NOW ready and unblocked
        newly_unblocked = current_ready_ids & prev_blocked_ids

        results: List[Dict[str, Any]] = []
        for row in self.get_ready_ops():
            oid = _safe_str(row.get("OP_ID")).upper()
            if oid in newly_unblocked:
                results.append(row)

        return results

    def save_snapshot(self) -> None:
        """Save current cache as previous snapshot for diff comparisons."""
        if self.CACHE_PATH.exists():
            import shutil
            prev = self.CACHE_PATH.with_suffix(".prev.json")
            try:
                shutil.copy2(str(self.CACHE_PATH), str(prev))
            except OSError:
                pass

    # ------------------------------------------------------------------
    # Cache metadata
    # ------------------------------------------------------------------

    def get_cache_stats(self) -> Dict[str, Any]:
        """Return metadata about the cache for diagnostics."""
        sheets = self._data.get("sheets", {})
        sheet_stats = {name: len(rows) for name, rows in sheets.items()}
        total_rows = sum(sheet_stats.values())

        return {
            "source": self._data.get("source", "NONE"),
            "generated_at": self._data.get("generated_at", "NONE"),
            "sheets_parsed": len(sheets),
            "total_rows": total_rows,
            "sheet_row_counts": sheet_stats,
            "cache_path": str(self.CACHE_PATH),
            "cache_exists": self.CACHE_PATH.exists(),
            "cache_stale": self._cache_stale(),
        }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _print_table(rows: List[Dict[str, Any]], columns: Optional[List[str]] = None,
                 max_col_width: int = 40) -> None:
    """Print rows as a formatted text table."""
    if not rows:
        print("  (no data)")
        return

    if columns is None:
        columns = list(rows[0].keys())

    # Truncate values for display
    def _trunc(val: Any) -> str:
        s = str(val) if val is not None else ""
        return s[:max_col_width] + ("..." if len(s) > max_col_width else "")

    # Column widths
    widths = {col: len(col) for col in columns}
    for row in rows[:200]:  # limit display
        for col in columns:
            widths[col] = max(widths[col], len(_trunc(row.get(col, ""))))

    # Header
    header = " | ".join(col.ljust(widths[col]) for col in columns)
    sep = "-+-".join("-" * widths[col] for col in columns)
    print("  {}".format(header))
    print("  {}".format(sep))

    # Rows
    for row in rows[:200]:
        line = " | ".join(_trunc(row.get(col, "")).ljust(widths[col]) for col in columns)
        print("  {}".format(line))

    if len(rows) > 200:
        print("  ... and {} more rows".format(len(rows) - 200))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Master Ops Bridge — Shared xlsx intelligence module",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --rebuild\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --stats\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --query CONTENT\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --venture C01\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --synergy\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --blockers\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --playbook C01\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --ready\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --json\n"
               "  python3 AUTOMATIONS/master_ops_bridge.py --brief CONTENT\n",
    )
    parser.add_argument("--rebuild", action="store_true", help="Force rebuild cache from xlsx")
    parser.add_argument("--stats", action="store_true", help="Print cache stats")
    parser.add_argument("--query", type=str, metavar="CATEGORY", help="Ops by category (CONTENT, SERVICE, APP, etc.)")
    parser.add_argument("--venture", type=str, metavar="ID", help="Venture details by ID")
    parser.add_argument("--op", type=str, metavar="OP_ID", help="Single op details by ID")
    parser.add_argument("--synergy", action="store_true", help="Show synergy stacks")
    parser.add_argument("--blockers", action="store_true", help="Show blocker summary")
    parser.add_argument("--playbook", type=str, metavar="OP_ID", help="Playbook for an op ID")
    parser.add_argument("--ready", action="store_true", help="Show READY ops")
    parser.add_argument("--actionable", action="store_true", help="Show daily actionable ops")
    parser.add_argument("--weights", type=str, metavar="OP_ID", help="Decision weights for an op")
    parser.add_argument("--brief", type=str, metavar="VENTURE_TYPE", help="Intelligence brief for venture type")
    parser.add_argument("--priority", action="store_true", help="Show priority launch list")
    parser.add_argument("--expansion", action="store_true", help="Show expansion queue")
    parser.add_argument("--infra", action="store_true", help="Show existing infrastructure")
    parser.add_argument("--tools", type=str, metavar="TYPE", help="Tool stack by type (video, hosting, lead_gen)")
    parser.add_argument("--alpha", action="store_true", help="Show alpha theses")
    parser.add_argument("--json", action="store_true", help="Full cache as JSON (for piping)")
    parser.add_argument("--snapshot", action="store_true", help="Save current cache as previous snapshot")
    args = parser.parse_args()

    # --rebuild first (or implicit if no cache)
    bridge = MasterOpsBridge(force_rebuild=args.rebuild)

    if args.json:
        json.dump(bridge._data, sys.stdout, indent=2, default=str)
        print()
        return

    if args.snapshot:
        bridge.save_snapshot()
        print("Snapshot saved.")
        return

    if args.stats:
        stats = bridge.get_cache_stats()
        print("\n  Master Ops Bridge — Cache Stats")
        print("  {}".format("=" * 45))
        print("  Source:       {}".format(stats["source"]))
        print("  Generated:    {}".format(stats["generated_at"]))
        print("  Sheets:       {}".format(stats["sheets_parsed"]))
        print("  Total rows:   {}".format(stats["total_rows"]))
        print("  Cache path:   {}".format(stats["cache_path"]))
        print("  Cache stale:  {}".format(stats["cache_stale"]))
        print("\n  Sheet breakdown:")
        for name, count in stats["sheet_row_counts"].items():
            print("    {:<30s} {:>5d} rows".format(name, count))
        print()
        return

    if args.query:
        ops = bridge.get_ops_by_category(args.query)
        print("\n  Ops in category '{}' ({}):\n".format(args.query, len(ops)))
        _print_table(ops, ["OP_ID", "OP_NAME", "REVENUE_RANGE", "AUTOMATION_LEVEL", "PRIORITY", "STATUS"])
        print()
        return

    if args.op:
        op = bridge.get_op(args.op)
        if op:
            print("\n  Op: {}\n".format(args.op))
            for k, v in op.items():
                print("    {}: {}".format(k, v))
            print()
        else:
            print("  Op '{}' not found.".format(args.op))
        return

    if args.venture:
        v = bridge.get_venture(args.venture)
        if v:
            print("\n  Venture: {}\n".format(args.venture))
            for k, val in v.items():
                print("    {}: {}".format(k, val))
            print()
        else:
            print("  Venture '{}' not found.".format(args.venture))
        return

    if args.synergy:
        stacks = bridge.get_synergy_stacks()
        print("\n  Synergy Stacks ({}):\n".format(len(stacks)))
        _print_table(stacks, ["PACKAGE_ID", "NAME", "SYNERGY_SCORE", "REVENUE_MULTIPLIER", "METHODS_COMBINED"])
        print()
        return

    if args.blockers:
        summary = bridge.get_blocker_summary()
        print("\n  Blocker Summary ({} unique blockers):\n".format(len(summary)))
        for b in summary:
            print("    {:<30s}  ops={}  ventures={}".format(
                b["blocker_key"], b["op_count"], b["venture_count"]))
            if b["blocked_ops"]:
                print("      ops: {}".format(", ".join(b["blocked_ops"][:10])))
            if b["blocked_ventures"]:
                print("      ventures: {}".format(", ".join(b["blocked_ventures"][:10])))
        print()
        return

    if args.playbook:
        steps = bridge.get_playbook_for_op(args.playbook)
        print("\n  Playbook for {} ({} steps):\n".format(args.playbook, len(steps)))
        _print_table(steps[:50], ["STEP_INDEX", "SECTION", "ROW_TYPE", "LANE", "TEXT"])
        print()
        return

    if args.ready:
        ops = bridge.get_ready_ops()
        print("\n  READY ops ({}):\n".format(len(ops)))
        _print_table(ops, ["OP_ID", "OP_NAME", "LANE", "AUTOMATION_SCORE_100", "SIGNAL_COUNT"])
        print()
        return

    if args.actionable:
        items = bridge.get_daily_actionable()
        print("\n  Daily Actionable ({}):\n".format(len(items)))
        _print_table(items, ["priority_rank", "op_id", "op_name", "lane", "automation_score", "next_action"])
        print()
        return

    if args.weights:
        w = bridge.get_decision_weights(args.weights)
        print("\n  Decision Weights for {}:\n".format(args.weights))
        for k, val in w.items():
            print("    {}: {}".format(k, val))
        print()
        return

    if args.brief:
        brief = bridge.get_ops_intelligence_brief(args.brief)
        print("\n  Intelligence Brief: {}".format(brief["venture_type"]))
        print("  {}".format("=" * 50))
        print("  Total ops:          {}".format(brief["ops_count"]))
        print("  Ready ops:          {}".format(brief["ready_count"]))
        print("  Blocked ops:        {}".format(brief["blocker_count"]))
        print("  Synergy combos:     {}".format(len(brief["synergies"])))
        print("  Alpha theses:       {}".format(len(brief["alpha_theses"])))
        print("  Playbook coverage:  {} ops".format(len(brief["playbook_available"])))
        print("  Priority launches:  {}".format(len(brief["priority_launches"])))
        if brief["ready_ops"]:
            print("\n  Ready ops:")
            for op in brief["ready_ops"][:10]:
                print("    {}: {}".format(_safe_str(op.get("OP_ID")), _safe_str(op.get("OP_NAME"))))
        if brief["synergies"]:
            print("\n  Synergies:")
            for s in brief["synergies"][:5]:
                print("    {}: {} (score={}, mult={})".format(
                    _safe_str(s.get("PACKAGE_ID")), _safe_str(s.get("NAME")),
                    _safe_str(s.get("SYNERGY_SCORE")), _safe_str(s.get("REVENUE_MULTIPLIER"))))
        if brief["alpha_theses"]:
            print("\n  Alpha theses:")
            for a in brief["alpha_theses"][:5]:
                print("    {}: {}".format(
                    _safe_str(a.get("ALPHA_ID")), _safe_str(a.get("OPPORTUNITY"))[:60]))
        if brief["blockers"]:
            print("\n  Blockers:")
            for b in brief["blockers"][:10]:
                print("    {}: {}".format(_safe_str(b.get("OP_ID")), _safe_str(b.get("BLOCKER_KEY"))))
        print()
        return

    if args.priority:
        items = bridge.get_priority_launch()
        print("\n  Priority Launch ({}):\n".format(len(items)))
        _print_table(items, ["RANK", "OP_ID", "OP_NAME", "WHY_NOW", "EFFORT", "TIME_TO_FIRST_$"])
        print()
        return

    if args.expansion:
        items = bridge.get_expansion_queue()
        print("\n  Expansion Queue ({}):\n".format(len(items)))
        _print_table(items, ["RANK", "OP_ID", "OP_NAME", "CATEGORY", "LANE", "EXPANSION_LOGIC"])
        print()
        return

    if args.infra:
        items = bridge.get_existing_infra()
        print("\n  Existing Infrastructure ({}):\n".format(len(items)))
        _print_table(items, ["CATEGORY", "ITEM", "FILE/LOCATION", "STATUS"])
        print()
        return

    if args.tools:
        items = bridge.get_tool_stack(args.tools)
        print("\n  Tool Stack: {} ({}):\n".format(args.tools, len(items)))
        if items:
            cols = list(items[0].keys())[:6]
            _print_table(items, cols)
        print()
        return

    if args.alpha:
        items = bridge.get_alpha_theses()
        print("\n  Alpha Theses ({}):\n".format(len(items)))
        _print_table(items, ["ALPHA_ID", "OPPORTUNITY", "LANE", "EDGE_DURATION", "WHY_LLM_EDGE"])
        print()
        return

    # Default: show stats + readiness summary
    stats = bridge.get_cache_stats()
    readiness = bridge.get_readiness_summary()
    print("\n  Master Ops Bridge")
    print("  {}".format("=" * 45))
    print("  Source: {} ({} rows across {} sheets)".format(
        stats["source"], stats["total_rows"], stats["sheets_parsed"]))
    print("  Generated: {}".format(stats["generated_at"]))
    print("  Readiness: {}".format(dict(readiness)))
    print("\n  Use --help for all options.\n")


if __name__ == "__main__":
    main()
